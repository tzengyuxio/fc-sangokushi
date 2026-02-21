#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
============================================================================
  FC/NES 三國志 (Sangokushi, KOEI 1988) — 武將資料解析與匯出工具 v2
============================================================================

【ROM 基本資訊】
  檔案格式 : iNES (.nes)
  Mapper   : 1 (MMC1)
  PRG ROM  : 256 KB (16 banks × 16 KB), 檔案偏移 0x00010 – 0x4000F
  CHR ROM  : 0 KB (使用 CHR-RAM, 圖形 tile 存放於 PRG ROM 中)

【武將資料表位置】
  所在 Bank : 14 (PRG bank 14, 檔案偏移 0x38010 – 0x3C00F)
  表頭      : 0x38010 – 0x38013, 共 4 bytes (4C 00 00 00)
  資料起始  : 0x38014
  記錄數    : 256 筆
  資料結束  : 0x39113

【單筆記錄結構】 — 每筆共 17 bytes

  ┌─────────────────── 12 bytes 武將資料 ───────────────────┐┌ 5 bytes 分隔符 ┐
  │ B0   B1   B2   B3   B4   B5   B6   B7   B8   B9  B10 B11│ 0A 0A 0A 00 00 │
  └─────────────────────────────────────────────────────────┘└────────────────┘

  Byte  偏移  欄位說明
  ───── ───── ──────────────────────────────────────────────
  B0    +0    年齡 (Age), signed byte
              正值 = 遊戲開始時的年齡 (0–65)
              負值 = 尚未出生 (絕對值為出生前年數, 例如 -17 = 17年後出生)
              已驗證：劉備=29, 關羽=28, 諸葛亮=8, 司馬師=-15, 劉禪=-17

  B1    +1    體力 (Body/Health), 範圍 15–100
              推測為武將壽命或健康度相關參數

  B2    +2    智力 (Intelligence), 範圍 15–100
              已驗證：劉備=95, 關羽=83, 諸葛亮=100

  B3    +3    武力 (Military/War), 範圍 15–100
              已驗證：劉備=63, 關羽=99, 諸葛亮=72

  B4    +4    魅力 (Charisma), 範圍 15–100
              已驗證：劉備=99, 關羽=70, 諸葛亮=97

  B5    +5    運氣/義理 (Luck/Fortune?), 範圍 15–100
              推測欄位, 尚未完全確認用途

  B6    +6    忠誠度 (Loyalty), 範圍 12–100
              君主固定為 100

  B7    +7    身份/水軍複合欄位 (Status + Navy), 值域 0–3
              此欄位以 bitfield 方式編碼兩項資訊:

              bit0 (0x01) = 水軍旗標 (Navy Flag)
                0 = 非水軍
                1 = 水軍
                已驗證水軍：曹操(1), 劉備(3), 關羽(43), 諸葛亮(174) 等
                已驗證非水軍：孫乾(47), 徐庶(173)

              bit1 (0x02) = 身份旗標 (Leader Flag)
                0 = 一般 (在野或武將, 依勢力歸屬判斷)
                1 = 統領 (君主或軍師)

              組合值:
                0 (0b00) = 一般・非水軍 (102人)
                1 (0b01) = 一般・水軍   (139人)
                2 (0b10) = 統領・非水軍 (3人, 軍師)
                3 (0b11) = 統領・水軍   (12人, 君主)

  B8-B9 +8,+9 初期兵士數 (Troops), Little-Endian 16-bit
              實際值 = B8 + B9 × 256
              已由使用者確認為兵士而非金錢

  B10   +10   城市位置 (City/Location ID), 範圍 0–55
              0 = 未配置

  B11   +11   勢力歸屬 (Faction ID), 範圍 0–14
              0 = 無勢力

  分隔符 +12–+16  固定為 0A 0A 0A 00 00 (5 bytes)

【勢力對照表】(根據記錄順序推斷)
  Faction  1 = 曹操    Faction  2 = 孫堅
  Faction  3 = 劉備    Faction  4 = 袁紹
  Faction  5 = 袁術    Faction  6 = 劉表
  Faction  7 = 董卓    Faction  8 = 劉焉
  Faction  9 = 馬騰    Faction 10 = 公孫瓚
  Faction 11 = 陶謙    Faction 12–14 = 其他
"""

import csv
import struct
import sys
import os
import warnings

# ─── 常數 ────────────────────────────────────────────────
TABLE_DATA_ADDR   = 0x38014
RECORD_DATA_SIZE  = 12
RECORD_SEP        = bytes([0x0A, 0x0A, 0x0A, 0x00, 0x00])
RECORD_TOTAL_SIZE = RECORD_DATA_SIZE + len(RECORD_SEP)  # 17
MAX_RECORDS       = 256

# 武將姓名表位置 (半角片假名)
NAME_TABLE_ADDR   = 0x3A314
NAME_RECORD_SIZE  = 15
NAME_DATA_SIZE    = 8  # 名字最多 8 bytes

# 漢字 tile ID 對照表 (從 ROM 姓名表 +8, +10, +12 位置解析)
KANJI_TILE_MAP = {
    0x01: "翊", 0x02: "苞", 0x03: "允", 0x04: "紘", 0x05: "羽", 0x06: "熙", 0x07: "叡",
    0x08: "蔡", 0x09: "永", 0x0A: "琦", 0x0B: "琮", 0x0C: "越", 0x0D: "延", 0x0E: "應",
    0x0F: "汜", 0x10: "王", 0x11: "黃", 0x12: "樊", 0x13: "稠", 0x14: "軫", 0x15: "夏",
    0x16: "華", 0x17: "旻", 0x18: "陶", 0x19: "璋", 0x1A: "沛", 0x1B: "郭", 0x1C: "懿",
    0x1D: "龐", 0x1E: "桓", 0x1F: "于", 0x20: "糜", 0x21: "桓", 0x22: "鮑", 0x23: "暹",
    0x24: "關", 0x25: "韓", 0x26: "玩", 0x27: "雍", 0x28: "曠", 0x29: "翔", 0x2A: "紀",
    0x2B: "儀", 0x2C: "宜", 0x2D: "休", 0x2E: "宮", 0x2F: "玠", 0x30: "許", 0x31: "顗",
    0x32: "興", 0x33: "欽", 0x34: "禁", 0x35: "瑾", 0x36: "金", 0x37: "傅", 0x38: "虞",
    0x39: "勳", 0x3A: "群", 0x3B: "邢", 0x3C: "珪", 0x3E: "堅", 0x3F: "憲", 0x40: "權",
    0x41: "謙", 0x42: "謙", 0x43: "嚴", 0x44: "玄", 0x45: "胡", 0x46: "顧", 0x47: "吳",
    0x49: "侯", 0x4A: "公", 0x4B: "孔", 0x4C: "洪", 0x4D: "晃", 0x4E: "洪", 0x4F: "紘",
    0x50: "高", 0x51: "綱", 0x52: "濟", 0x53: "策", 0x54: "索", 0x55: "司", 0x56: "史",
    0x57: "士", 0x58: "師", 0x59: "志", 0x5A: "慈", 0x5B: "治", 0x5C: "竺", 0x5D: "芝",
    0x5E: "朱", 0x5F: "儒", 0x60: "授", 0x62: "周", 0x64: "脩", 0x65: "脩", 0x66: "繡",
    0x67: "醜", 0x68: "肅", 0x69: "術", 0x6A: "循", 0x6B: "荀", 0x6C: "純", 0x6D: "庶",
    0x6E: "諸", 0x6F: "徐", 0x70: "紹", 0x71: "昭", 0x72: "昭", 0x73: "紹", 0x74: "蔣",
    0x75: "鍾", 0x76: "植", 0x77: "信", 0x78: "審", 0x79: "真", 0x7A: "辛", 0x7B: "進",
    0x7C: "仁", 0x7D: "圖", 0x7E: "遂", 0x7F: "成", 0x80: "正", 0x81: "盛", 0x83: "籍",
    0x84: "績", 0x85: "旋", 0x86: "選", 0x87: "全", 0x88: "禪", 0x89: "祖", 0x8A: "雙",
    0x8B: "倉", 0x8C: "宋", 0x8D: "操", 0x8E: "曹", 0x91: "孫", 0x92: "遜", 0x93: "太",
    0x94: "岱", 0x95: "泰", 0x97: "卓", 0x98: "澤", 0x99: "達", 0x9A: "堪", 0x9B: "中",
    0x9C: "忠", 0x9D: "丁", 0x9E: "寵", 0x9F: "張", 0xA0: "超", 0xA1: "陳", 0xA2: "定",
    0xA3: "程", 0xA4: "鐵", 0xA5: "典", 0xA7: "登", 0xA8: "度", 0xA9: "當", 0xAA: "統",
    0xAB: "董", 0xAC: "陶", 0xAD: "騰", 0xAE: "銅", 0xAF: "道", 0xB0: "德", 0xB1: "惇",
    0xB2: "任", 0xB3: "寧", 0xB4: "之", 0xB5: "巴", 0xB6: "馬", 0xB7: "配", 0xB8: "薄",
    0xB9: "班", 0xBA: "範", 0xBB: "費", 0xBC: "飛", 0xBD: "備", 0xBE: "彪", 0xBF: "表",
    0xC0: "豹", 0xC1: "布", 0xC2: "普", 0xC3: "武", 0xC4: "封", 0xC5: "淵", 0xC6: "文",
    0xC7: "平", 0xC8: "圃", 0xC9: "步", 0xCA: "奉", 0xCB: "法", 0xCC: "芳", 0xCD: "褒",
    0xCE: "豐", 0xCF: "翻", 0xD0: "摩", 0xD1: "滿", 0xD2: "孟", 0xD3: "毛", 0xD4: "蒙",
    0xD5: "靖", 0xD6: "優", 0xD8: "雄", 0xD9: "融", 0xDC: "楊", 0xDD: "雷", 0xDE: "蘭",
    0xDF: "覽", 0xE0: "李", 0xE1: "理", 0xE2: "陸", 0xE3: "劉", 0xE4: "隆", 0xE5: "亮",
    0xE6: "凌", 0xE7: "梁", 0xE9: "良", 0xEA: "遼", 0xEB: "琳", 0xEC: "累", 0xEE: "靈",
    0xEF: "呂", 0xF0: "魯", 0xF1: "朗", 0xF2: "和", 0xF6: "袁", 0xF7: "焉", 0xF8: "瓚",
    0xF9: "荀", 0xFA: "彧", 0xFB: "昱", 0xFC: "韋", 0xFD: "曄", 0xFE: "攸", 0xFF: "丕",
}

# ─── 外部提供資料 (非 ROM 解析) ─────────────────────────
# 標記 [EXT] 表示此資料來自使用者提供, 非從 ROM 中直接讀取
# 格式: 序號 → (姓名, 假名)
# 此為靜態備用資料，優先使用動態載入的外部 CSV 資料
EXT_CHAR_INFO = {
    1:  ("曹操", "そうそう"),
    2:  ("孫堅", "そんけん"),
    3:  ("劉備", "りゅうび"),
    4:  ("袁紹", "えんしょう"),
    5:  ("袁術", "えんしゅう"),
    6:  ("劉表", "りゅうびょう"),
    7:  ("董卓", "とうたく"),
    8:  ("劉焉", "りゅうえん"),
    9:  ("馬騰", ""),
    10: ("公孫瓚", "こうそんさん"),
    43: ("關羽", "かんう"),
    44: ("張飛", "ちょうひ"),
    47: ("孫亁", "そんかん"),
    173: ("徐庶", ""),
    174: ("諸葛亮", "しょかつりょう"),
    230: ("孟獲", "もうかく"),
    231: ("雍闓", "よんがい"),
}

# 外部 CSV 檔案路徑 (光榮三國志系列武將登場統計)
EXT_CSV_PATH = "光榮三國志系列武將登場統計 - 能力表.csv"


def load_ext_char_info_from_csv(csv_path):
    """
    從外部 CSV 載入 S01 (FC 三國志) 武將資料

    回傳: {(body, intelligence, military, charisma, luck): (name, kana)} 字典

    CSV 欄位對應:
      - 欄 0: 姓名
      - 欄 5: 假名 (日文假名讀音)
      - 欄 8-12: S01 身體/知力/武力/魅力/運勢
    """
    if not os.path.exists(csv_path):
        return {}

    lookup = {}
    try:
        with open(csv_path, "r", encoding="utf-8") as f:
            reader = csv.reader(f)
            next(reader)  # 跳過表頭
            for row in reader:
                if len(row) <= 12 or not row[8]:
                    continue  # 跳過無 S01 資料的列
                try:
                    body = int(row[8])
                    intl = int(row[9])
                    mil = int(row[10])
                    cha = int(row[11])
                    luck = int(row[12])
                except ValueError:
                    continue  # 跳過能力值欄位為空或非數字的列

                key = (body, intl, mil, cha, luck)
                name = row[0]
                kana = row[5]

                if key in lookup:
                    # 多筆武將能力值相同，記錄警告
                    existing_name = lookup[key][0]
                    warnings.warn(
                        f"重複能力值組合 {key}: '{existing_name}' 與 '{name}'，採用第一筆",
                        stacklevel=2
                    )
                else:
                    lookup[key] = (name, kana)
    except Exception as e:
        warnings.warn(f"載入外部 CSV 失敗: {e}", stacklevel=2)
        return {}

    return lookup

# 身份名稱
def get_status_name(b7):
    return {0: "一般・非水軍", 1: "一般・水軍", 2: "統領・非水軍", 3: "統領・水軍"}.get(b7, f"未知({b7})")

def get_role_name(b7):
    """從 B7 bit1 取得身份角色"""
    return "統領" if (b7 >> 1) & 1 else "一般"

def get_navy(b7):
    """從 B7 bit0 取得水軍旗標"""
    return b7 & 1


def decode_halfwidth_kana(data):
    """
    解碼半角片假名 (Shift-JIS 0xA6-0xDF)

    參數:
        data: bytes, 最多 8 bytes 的假名資料

    回傳:
        str: 解碼後的假名字串
    """
    result = []
    for b in data:
        if b == 0x00:
            break  # 名字結束
        if 0xA6 <= b <= 0xDF:
            # 半角片假名範圍
            result.append(bytes([b]).decode('cp932', errors='replace'))
        # 忽略其他字元
    return ''.join(result)


def decode_kanji_tiles(data):
    """
    從姓名記錄的 metadata 解碼漢字 (tile ID → 漢字)

    參數:
        data: bytes, 姓名記錄 +8 到 +14 的 7 bytes

    回傳:
        str: 解碼後的漢字字串 (最多 3 字)

    姓名記錄結構:
        +0-7:  假名 (8 bytes)
        +8:    第一個漢字 tile ID
        +9:    未知
        +10:   第二個漢字 tile ID
        +11:   未知
        +12:   第三個漢字 tile ID
        +13-14: 未知
    """
    kanji_chars = []
    # tile ID 位於 offset 0, 2, 4 (相對於 +8 起始)
    for offset in [0, 2, 4]:
        if offset < len(data):
            tile_id = data[offset]
            if tile_id in KANJI_TILE_MAP:
                kanji_chars.append(KANJI_TILE_MAP[tile_id])
    return ''.join(kanji_chars)


def load_rom_names(rom_data):
    """
    從 ROM 載入武將姓名表 (假名與漢字)

    參數:
        rom_data: bytes, ROM 資料

    回傳:
        list of tuples: [(kana, kanji), ...] 共 257 筆 (索引 0 為新君主模板)
    """
    names = []
    for i in range(257):  # 256 武將 + 1 新君主模板
        offset = NAME_TABLE_ADDR + i * NAME_RECORD_SIZE
        if offset + NAME_RECORD_SIZE > len(rom_data):
            break
        # 假名 (前 8 bytes)
        kana_bytes = rom_data[offset:offset + NAME_DATA_SIZE]
        kana = decode_halfwidth_kana(kana_bytes)
        # 漢字 tile IDs (後 7 bytes: +8 到 +14)
        kanji_bytes = rom_data[offset + NAME_DATA_SIZE:offset + NAME_RECORD_SIZE]
        kanji = decode_kanji_tiles(kanji_bytes)
        names.append((kana, kanji))
    return names

# ─── CSV 欄位 ────────────────────────────────────────────
CSV_FIELDS = [
    "Index",              # 記錄序號
    "ROM_Offset",         # 檔案內偏移 (hex)
    "ROM_Kana",           # 假名 (ROM 解析, 半角片假名)
    "ROM_Kanji",          # 漢字 (ROM 解析, tile ID 對照)
    "[EXT]Name",          # 姓名 (外部提供)
    "[EXT]Kana",          # 假名 (外部提供)
    "Age",                # B0 - 年齡 (signed)
    "Body",               # B1 - 體力
    "Intelligence",       # B2 - 智力
    "Military",           # B3 - 武力
    "Charisma",           # B4 - 魅力
    "Luck",               # B5 - 運氣 (推測)
    "Loyalty",            # B6 - 忠誠
    "B7_Raw",             # B7 - 原始值
    "Navy",               # B7 bit0 - 水軍
    "Role",               # B7 bit1 - 身份
    "Troops",             # B8-B9 - 兵士數
    "City",               # B10 - 城市
    "Faction",            # B11 - 勢力
    "Raw_Hex",            # 原始 12 bytes
]


def parse_record(rom_data, offset):
    if offset + RECORD_TOTAL_SIZE > len(rom_data):
        return None
    rec = rom_data[offset:offset + RECORD_DATA_SIZE]
    sep = rom_data[offset + RECORD_DATA_SIZE:offset + RECORD_TOTAL_SIZE]
    if sep != RECORD_SEP:
        return None
    age = struct.unpack('b', bytes([rec[0]]))[0]
    return {
        "age":          age,
        "body":         rec[1],
        "intelligence": rec[2],
        "military":     rec[3],
        "charisma":     rec[4],
        "luck":         rec[5],
        "loyalty":      rec[6],
        "b7_raw":       rec[7],
        "navy":         get_navy(rec[7]),
        "role":         get_role_name(rec[7]),
        "troops":       rec[8] + rec[9] * 256,
        "city":         rec[10],
        "faction":      rec[11],
        "raw":          rec,
    }


def extract_all(rom_path, ext_csv_path=None):
    """
    從 ROM 檔案解析所有武將記錄

    參數:
        rom_path: ROM 檔案路徑
        ext_csv_path: 外部 CSV 路徑 (可選，用於動態載入武將姓名)

    回傳:
        武將記錄列表
    """
    with open(rom_path, "rb") as f:
        rom = f.read()
    if rom[:4] != b"NES\x1a":
        raise ValueError("非有效的 iNES ROM 檔案")

    # 載入外部 CSV 資料 (依能力值查找姓名)
    ext_lookup = {}
    if ext_csv_path:
        ext_lookup = load_ext_char_info_from_csv(ext_csv_path)

    # 載入 ROM 中的假名姓名表
    rom_names = load_rom_names(rom)

    records = []
    offset = TABLE_DATA_ADDR
    for idx in range(MAX_RECORDS):
        rec = parse_record(rom, offset)
        if rec is None:
            break
        rec["index"] = idx
        rec["offset"] = offset

        # ROM 假名與漢字 (名字表索引與武將索引相同)
        if idx < len(rom_names):
            rec["rom_kana"] = rom_names[idx][0]
            rec["rom_kanji"] = rom_names[idx][1]
        else:
            rec["rom_kana"] = ""
            rec["rom_kanji"] = ""

        # 先嘗試從外部 CSV 比對能力值
        stats_key = (rec["body"], rec["intelligence"], rec["military"],
                     rec["charisma"], rec["luck"])
        if stats_key in ext_lookup:
            name, kana = ext_lookup[stats_key]
        else:
            # 回退到靜態 EXT_CHAR_INFO (依序號)
            name, kana = EXT_CHAR_INFO.get(idx, ("", ""))

        rec["ext_name"] = name
        rec["ext_kana"] = kana
        records.append(rec)
        offset += RECORD_TOTAL_SIZE
    return records


def export_csv(records, output_path):
    with open(output_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=CSV_FIELDS)
        writer.writeheader()
        for r in records:
            writer.writerow({
                "Index":        r["index"],
                "ROM_Offset":   f'0x{r["offset"]:05X}',
                "ROM_Kana":     r["rom_kana"],
                "ROM_Kanji":    r["rom_kanji"],
                "[EXT]Name":    r["ext_name"],
                "[EXT]Kana":    r["ext_kana"],
                "Age":          r["age"],
                "Body":         r["body"],
                "Intelligence": r["intelligence"],
                "Military":     r["military"],
                "Charisma":     r["charisma"],
                "Luck":         r["luck"],
                "Loyalty":      r["loyalty"],
                "B7_Raw":       r["b7_raw"],
                "Navy":         "水軍" if r["navy"] else "",
                "Role":         r["role"],
                "Troops":       r["troops"],
                "City":         r["city"],
                "Faction":      r["faction"],
                "Raw_Hex":      " ".join(f"{b:02X}" for b in r["raw"]),
            })
    print(f"已匯出 {len(records)} 筆 → {output_path}")


def export_xlsx(records, output_path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "武將資料"

    headers = [
        ("序號", 5),  ("偏移", 9),   ("ROM假名", 12),  ("ROM漢字", 10),
        ("[EXT]姓名", 10), ("[EXT]假名", 14),
        ("年齡", 5),  ("體力", 5),   ("智力", 5),       ("武力", 5),
        ("魅力", 5),  ("運氣", 5),   ("忠誠", 5),       ("B7", 4),
        ("水軍", 5),  ("身份", 10),  ("兵士", 8),       ("城市", 5),
        ("勢力", 5),  ("Raw Hex", 38),
    ]

    hdr_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    hdr_fill = PatternFill("solid", fgColor="4472C4")
    ext_fill = PatternFill("solid", fgColor="2E75B6")  # 外部欄位用深藍
    hdr_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(bottom=Side(style="thin", color="D0D0D0"))

    for col, (name, width) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.font = hdr_font
        cell.fill = ext_fill if name.startswith("[EXT]") else hdr_fill
        cell.alignment = hdr_align
        ws.column_dimensions[cell.column_letter].width = width

    # 條件色
    leader_fill = PatternFill("solid", fgColor="FFF2CC")
    navy_fill   = PatternFill("solid", fgColor="D6EAF8")
    both_fill   = PatternFill("solid", fgColor="D5F5E3")
    data_font   = Font(name="Arial", size=10)
    mono_font   = Font(name="Consolas", size=9, color="666666")
    ext_font    = Font(name="Arial", size=10, color="1A5276")
    center      = Alignment(horizontal="center")
    neg_font    = Font(name="Arial", size=10, color="CC0000")

    for i, r in enumerate(records):
        row = i + 2
        is_navy   = r["navy"]
        is_leader = (r["b7_raw"] >> 1) & 1
        bg = None
        if is_leader and is_navy:
            bg = both_fill
        elif is_leader:
            bg = leader_fill
        elif is_navy:
            bg = navy_fill

        values = [
            r["index"],
            f'0x{r["offset"]:05X}',
            r["rom_kana"],
            r["rom_kanji"],
            r["ext_name"],
            r["ext_kana"],
            r["age"],
            r["body"],
            r["intelligence"],
            r["military"],
            r["charisma"],
            r["luck"],
            r["loyalty"],
            r["b7_raw"],
            "●" if r["navy"] else "",
            r["role"],
            r["troops"],
            r["city"],
            r["faction"],
            " ".join(f"{b:02X}" for b in r["raw"]),
        ]

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = thin_border
            cell.alignment = center if col not in (3, 4, 5, 6, 20) else Alignment(horizontal="left")
            # 字型
            if col == 20:  # Raw Hex
                cell.font = mono_font
            elif col in (3, 4):  # ROM假名, ROM漢字
                cell.font = data_font
            elif col in (5, 6):  # [EXT]姓名, [EXT]假名
                cell.font = ext_font
            elif col == 7 and isinstance(val, int) and val < 0:  # 年齡 (負值)
                cell.font = neg_font
            else:
                cell.font = data_font
            # 背景
            if bg:
                cell.fill = bg

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(output_path)
    print(f"已匯出 {len(records)} 筆 → {output_path}")


if __name__ == "__main__":
    rom_path = sys.argv[1] if len(sys.argv) > 1 else "Sangokushi__Japan_.nes"

    if not os.path.exists(rom_path):
        print(f"錯誤: 找不到 ROM 檔案 '{rom_path}'")
        print(f"用法: python {sys.argv[0]} <rom_file.nes>")
        sys.exit(1)

    # 決定外部 CSV 路徑 (同目錄下尋找)
    script_dir = os.path.dirname(os.path.abspath(__file__))
    ext_csv_path = os.path.join(script_dir, EXT_CSV_PATH)
    if not os.path.exists(ext_csv_path):
        ext_csv_path = None
        print(f"提示: 未找到外部 CSV '{EXT_CSV_PATH}'，使用靜態姓名資料")

    records = extract_all(rom_path, ext_csv_path)

    navy_count   = sum(1 for r in records if r["navy"])
    leader_count = sum(1 for r in records if (r["b7_raw"] >> 1) & 1)
    named_count  = sum(1 for r in records if r["ext_name"])

    print(f"解析完成: 共 {len(records)} 筆武將記錄")
    print(f"  水軍: {navy_count} 人 / 非水軍: {len(records) - navy_count} 人")
    print(f"  統領: {leader_count} 人 (君主+軍師)")
    print(f"  已知姓名: {named_count} 人 (外部提供)")

    base = os.path.splitext(os.path.basename(rom_path))[0]
    export_csv(records, f"{base}_characters_v2.csv")

    try:
        import openpyxl
        export_xlsx(records, f"{base}_characters_v2.xlsx")
    except ImportError:
        print("提示: 安裝 openpyxl 可額外匯出 .xlsx")
